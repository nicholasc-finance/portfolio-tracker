import csv
import os
from tabulate import tabulate
import yfinance as yf
import time
from colorama import Fore, Style, init
init(autoreset=True)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import PieChart, BarChart, Reference, Series
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from openpyxl.chart.label import DataLabelList

def get_valid_number(prompt, number_type=float):
    while True:
        try:
            return number_type(input(prompt))
        except ValueError:
            print("⚠️  Invalid input. Please enter a valid number.")

portfolio = []
CSV_FILENAME = 'portfolio_tracker.csv'

def fetch_latest_price(ticker):
    time.sleep(0.5)
    try:
        data = yf.Ticker(ticker).history(period='1d')
        if not data.empty:
            return data['Close'].iloc[-1]
    except Exception as e:
        print(f"⚠️  Error fetching price for {ticker}: {e}")
        return None
    
def calculate_stock_metrics(stock):
    stock['value'] = stock['price'] * stock['shares']
    stock['gain_loss'] = (stock['price'] - stock['buy_price']) * stock['shares']
    stock['gain_loss_percent'] = ((stock['price'] - stock['buy_price']) / stock['buy_price']) * 100
    return stock

def load_portfolio_from_csv(filename):
    loaded_portfolio = []
    if os.path.exists(filename):
        with open(filename, mode='r', encoding='utf-8') as file:
            reader = csv.DictReader(file)
            for row in reader:
                stock = {
                    'name': row['Stocks'],
                    'ticker': row['Ticker'],
                    'shares': int(row['Shares']),
                    'buy_price': float(row['Buy Price (RM)'])
                }
                loaded_portfolio.append(stock)
    return loaded_portfolio

def save_portfolio_to_csv(filename, portfolio):
    with open(filename, mode='w', newline='', encoding='utf-8') as file:
        fieldnames = [
            'Stocks', 'Ticker', 'Shares',
            'Buy Price (RM)', 'Current Price (RM)',
            'Value (RM)', 'Gain/Loss (RM)', 'Gain/Loss (%)'
        ]
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()   
        for s in portfolio:
            if 'price' not in s or 'value' not in s:
                latest_price = fetch_latest_price(s['ticker'])
                s['price'] = latest_price if latest_price else s.get('price', 0)
                s = calculate_stock_metrics(s)

            writer.writerow({
                'Stocks': s['name'],
                'Ticker': s['ticker'],
                'Shares': s['shares'],
                'Buy Price (RM)': f"{s['buy_price']:.3f}",
                'Current Price (RM)': f"{s.get('price', 0):.3f}",
                'Value (RM)': f"{s['value']:.2f}",
                'Gain/Loss (RM)': f"{s['gain_loss']:.2f}",
                'Gain/Loss (%)': f"{s['gain_loss_percent']:.2f}"
            })

def export_to_excel(portfolio, filename='portfolio_dashboard.xlsx'):
    if not portfolio:
        print('⚠️  Portfolio is empty. Nothing to export.')
        return
    
    for s in portfolio:
        if 'price' not in s or s['price'] == 0:
            latest_price = fetch_latest_price(s['ticker'])
            s['price'] = latest_price if latest_price else 0
        calculate_stock_metrics(s)
    
    if os.path.exists(filename):
        print(f"⚠️  File '{filename}' already exists.")
        choice = input("Choose: (1) Overwrite, (2) Create new file, (3) Cancel: ")
        
        if choice == '1':
            try:
                os.remove(filename)
                print("🔄 Overwriting existing file...")
            except PermissionError:
                print("❌ Cannot overwrite - file is open in Excel or another program.")
                return
        elif choice == '2':
            base_name = os.path.splitext(filename)[0]
            counter = 1
            while True:
                new_filename = f"{base_name}_{counter}.xlsx"
                if not os.path.exists(new_filename):
                    filename = new_filename
                    break
                counter += 1
            print(f"📝 Creating new file: {filename}")
        elif choice == '3':
            print("❌ Export cancelled.")
            return
        else:
            print("❌ Invalid choice. Export cancelled.")
            return
        
    wb = Workbook()
    ws = wb.active
    ws.title = "Portfolio Dashboard"

    total_value = sum(s['value'] for s in portfolio)
    total_gain = sum(s['gain_loss'] for s in portfolio)

    ws.merge_cells('A1:H1') 
    ws['A1'] = "Portfolio Summary"
    ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    summary_data = [
        ["Total Portfolio Value (RM)", total_value],
        ["Total Gain/Loss (RM)", total_gain],
        ["Number of Stocks", len(portfolio)]
    ]
    
    for i, (label, value) in enumerate(summary_data, start=3):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value
        ws[f'A{i}'].font = Font(bold=True)
        ws[f'A{i}'].alignment = Alignment(horizontal='left')
        ws[f'B{i}'].font = Font(bold=True)
        if isinstance(value, (int, float)):
            ws[f'B{i}'].number_format = '#,##0.00'

    headers = ['Stocks', 'Ticker', 'Shares', 'Buy Price (RM)', 'Current Price (RM)',
               'Value (RM)', 'Gain/Loss (RM)', 'Gain/Loss (%)']
    
    table_start_row = 8
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=table_start_row, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    for row_idx, stock in enumerate(portfolio, start=table_start_row + 1):
        ws.cell(row=row_idx, column=1, value=stock['name'])
        ws.cell(row=row_idx, column=2, value=stock['ticker'])
        ws.cell(row=row_idx, column=3, value=stock['shares'])
        ws.cell(row=row_idx, column=4, value=stock['buy_price'])
        ws.cell(row=row_idx, column=5, value=stock['price'])
        ws.cell(row=row_idx, column=6, value=stock['value'])
        ws.cell(row=row_idx, column=7, value=stock['gain_loss'])
        ws.cell(row=row_idx, column=8, value=stock['gain_loss_percent'])

    number_columns = ['D', 'E', 'F', 'G', 'H']
    for col in number_columns:
        for row in range(table_start_row + 1, table_start_row + len(portfolio) + 1):
            cell = ws[f'{col}{row}']
            if col in ['D', 'E']: 
                cell.number_format = '0.000'
            elif col in ['F', 'G']:
                cell.number_format = '#,##0.00'
            elif col == 'H':
                cell.number_format = '0.00'

    for col in ws.columns:
        col_cells = [cell for cell in col if not isinstance(cell, MergedCell)]
        if not col_cells:
            continue
        column_letter = col_cells[0].column_letter
        max_length = 0
        for cell in col_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    gain_loss_col = 'G'
    last_row = table_start_row + len(portfolio)
    ws.conditional_formatting.add(
        f'{gain_loss_col}{table_start_row + 1}:{gain_loss_col}{last_row}',
        CellIsRule(operator='greaterThan', formula=['0'], 
                  fill=PatternFill(start_color='C6EFCE', fill_type='solid'))
    )
    ws.conditional_formatting.add(
        f'{gain_loss_col}{table_start_row + 1}:{gain_loss_col}{last_row}',
        CellIsRule(operator='lessThan', formula=['0'], 
                  fill=PatternFill(start_color='FFC7CE', fill_type='solid'))
    )

    if portfolio:
        chart_start_row = last_row + 3

        pie = PieChart()
        pie.title = "Portfolio Composition by Value"

        labels = Reference(ws, min_col=1, min_row=table_start_row + 1, max_row=last_row)
        data = Reference(ws, min_col=6, min_row=table_start_row + 1, max_row=last_row)
        series = Series(data, title="Portfolio %")
        pie.series.append(series)
        pie.set_categories(labels)

        pie.dataLabels = DataLabelList()
        pie.dataLabels.showCategoryName = True
        pie.dataLabels.showPercent = True
        pie.dataLabels.showLeaderLines = True
        pie.dataLabels.showSeriesName = False
        pie.dataLabels.showVal = False

        pie.height = 12
        pie.width = 14
        ws.add_chart(pie, f"A{chart_start_row}")

        bar = BarChart()
        bar.title = "Gain/Loss by Stock"
        bar.y_axis.title = "Amount (RM)"
        bar.x_axis.title = "Stocks"

        bar_labels = Reference(ws, min_col=1, min_row=table_start_row + 1, max_row=last_row)
        bar_data = Reference(ws, min_col=7, min_row=table_start_row + 1, max_row=last_row)
        bar_series = Series(bar_data, title="Gain/Loss (RM)")
        bar.series.append(bar_series)
        bar.set_categories(bar_labels)

        bar.dataLabels = DataLabelList()
        bar.dataLabels.showVal = True
        bar.dataLabels.showCategoryName = True
        bar.dataLabels.showSeriesName = False

        bar.height = 12
        bar.width = 14
        ws.add_chart(bar, f"J{chart_start_row}")

    wb.save(filename)
    time.sleep(0.5)
    print(f"✅ Portfolio successfully exported to {filename}")
    print("📊 You can now open your Excel dashboard to view the summary.")

def add_new_stock(portfolio):
    num_new = get_valid_number('How many new stocks do you want to add? ', int)
    for i in range(num_new):
        print(f'\nAdding Stock {i+1}')
        stock = {}
        stock['ticker'] = input('Enter stock ticker (e.g. 4707.KL for NESTLE): ').upper().strip()
        stock['name'] = input('Enter stock name: ').strip()
        stock['shares'] = get_valid_number('Enter number of shares: ', int)
        stock['buy_price'] = get_valid_number('Enter your purchase price per share (RM): ')

        latest_price = fetch_latest_price(stock['ticker'])
        if latest_price:
            stock['price'] = latest_price
            print(f"Fetched latest price: RM{stock['price']:.3f}")
        else:
            print("⚠️  No data found. Please enter price manually.")
            stock['price'] = get_valid_number("Enter stock price manually: ")
        
        stock = calculate_stock_metrics(stock)
        portfolio.append(stock)
        print(f"✅ Added {stock['name']} ({stock['ticker']}) successfully.")

def remove_stock(portfolio):
    if not portfolio:
        print("⚠️  Portfolio is empty.")
        return
    
    print("\n🗑️  Remove a Stock from Portfolio")
    ticker = input("Enter the ticker symbol of the stock to remove: ").upper()

    for stock in portfolio:
        if stock['ticker'] == ticker:
            portfolio.remove(stock)
            print(f"✅ {stock['name']} ({ticker}) has been removed.")
            save_portfolio_to_csv(CSV_FILENAME, portfolio)
            return
        
    print(f"⚠️  Stock with ticker '{ticker}' not found in portfolio.")

def edit_stock(portfolio):
    if not portfolio:
        print("⚠️  Portfolio is empty.")
        return
    
    print("\n📝 Edit a Stock in Portfolio")
    ticker = input("Enter the ticker symbol of the stock to edit: ").upper()

    for stock in portfolio:
        if stock['ticker'] == ticker:
            print(f"\nEditing {stock['name']} ({ticker})")
            print("1. Edit Stock Name")
            print("2. Edit Number of Shares")
            print("3. Edit Buy Price (RM)")
            choice = input("Enter your choice (1-3): ")

            if choice == '1':
                stock['name'] = input("Enter new stock name: ")
            elif choice == '2':
                stock['shares'] = get_valid_number("Enter new number of shares: ", int)
            elif choice == '3':
                stock['buy_price'] = get_valid_number("Enter new buy price (RM): ")
            else:
                print("⚠️  Invalid choice.")
                return
            
            latest_price = fetch_latest_price(stock['ticker'])
            stock['price'] = latest_price if latest_price else stock.get('price', 0)
            stock = calculate_stock_metrics(stock)

            print(f"✅ {stock['name']} ({ticker}) has been updated successfully!")
            save_portfolio_to_csv(CSV_FILENAME, portfolio)
            return
    
    print(f"⚠️  Stock with ticker '{ticker}' not found in portfolio.")

def view_portfolio_summary(portfolio):
    if not portfolio:
        print("⚠️  Portfolio is empty.")
        return
    
    total_value = 0
    total_gain = 0

    for stock in portfolio:
        latest_price = fetch_latest_price(stock['ticker'])
        if latest_price:
            stock['price'] = latest_price
        else:
            print(f"⚠️  Could not fetch price for {stock['ticker']}. Using previous price.")
        stock = calculate_stock_metrics(stock)
        total_value += stock['value']
        total_gain += stock['gain_loss']

    headers = ['Stocks', 'Ticker', 'Shares', 'Buy Price (RM)', 'Current Price (RM)', 'Value (RM)', 'Gain/Loss (RM)', 'Gain/Loss (%)']
    table_data = []
    for s in portfolio:
        gain_color = (
            Fore.GREEN if s['gain_loss'] > 0
            else Fore.RED if s['gain_loss'] < 0
            else Fore.YELLOW
        )

        table_data.append([
            s['name'],
            s['ticker'],
            s['shares'],
            f"{s['buy_price']:.3f}",
            f"{s['price']:.3f}",
            f"{s['value']:.2f}",
            f"{gain_color}{s['gain_loss']:.2f}{Style.RESET_ALL}",
            f"{gain_color}{s['gain_loss_percent']:.2f}{Style.RESET_ALL}"

        ])
        
    

    print("\n📊 Portfolio Summary:")
    print(tabulate(table_data, headers=headers, tablefmt='fancy_grid'))
    print(f"\n💰 Total Portfolio Value: RM {total_value:.2f}")
    print(f"📈 Total Gain/Loss: RM {total_gain:.2f}")

    save_portfolio_to_csv(CSV_FILENAME, portfolio)
    print("💾 Portfolio saved successfully.")

def refresh_prices(portfolio):
    if not portfolio:
        print("⚠️  Portfolio is empty.")
        return
    
    print("\n🔄 Refreshing all stock prices...")
    for stock in portfolio:
        latest_price = fetch_latest_price(stock['ticker'])
        time.sleep(0.1)
        if latest_price:
            stock['price'] = latest_price
            stock = calculate_stock_metrics(stock)
            print(f"✅ Updated {stock['name']} ({stock['ticker']}) → RM{stock['price']:.3f}")
        else:
            print(f"⚠️  Could not fetch price for {stock['ticker']}.")
    
    save_portfolio_to_csv(CSV_FILENAME, portfolio)
    time.sleep(0.3)
    print("\n💾 All stock prices refreshed and saved successfully!")

def main():
    portfolio = load_portfolio_from_csv(CSV_FILENAME)
    
    while True:
        print("\n====== Portfolio Management Menu ======")
        print("1. View Portfolio Summary")
        print("2. Add new stock")
        print("3. Remove stock")
        print("4. Edit stock")
        print("5. Refresh stock prices")
        print("6. Export to Excel")
        print("7. Exit")
        print("=======================================")

        choice = input("Enter your choice (1-7): ")  

        if choice == '1':
            print("\n[Viewing Portfolio Summary...]\n")
            time.sleep(0.2)
            view_portfolio_summary(portfolio)
        
        elif choice == '2':
            print("\n[Adding new stock...]\n")
            time.sleep(0.5)
            add_new_stock(portfolio)
            save_portfolio_to_csv(CSV_FILENAME, portfolio)

        elif choice == '3':
            print("\n[Removing stock...]\n")
            time.sleep(0.5)
            remove_stock(portfolio)
        
        elif choice == '4':
            print("\n[Editing stock...]\n")
            time.sleep(0.5)
            edit_stock(portfolio)
        
        elif choice == '5':
            print("\n[Refreshing stock prices...]\n")
            time.sleep(0.5)
            refresh_prices(portfolio)
        
        elif choice == '6':
            print("\n[Exporting portfolio to Excel...]\n")
            time.sleep(0.5)
            export_to_excel(portfolio)

        elif choice == '7':
            time.sleep(0.7)
            print("\nExiting program. Goodbye!")
            break

        else:
            print("⚠️  Invalid choice. Please enter 1, 2, or 3.")

if __name__ == "__main__":
    main()